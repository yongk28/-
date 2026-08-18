"""
DART 전체법인 + OpenDART 기업 스크리너 - Streamlit 단일 페이지 앱

실행 방법:
    pip install streamlit pandas requests openpyxl
    streamlit run app.py

배포 방법 (무료로 URL 받기):
    1. 이 파일을 GitHub 저장소에 올린다
    2. share.streamlit.io 에서 저장소 연결 -> 자동 배포
    (API 키는 화면에서 매번 입력하는 방식 - 코드에 하드코딩하지 말 것)
"""

import io
import re
import time
import threading
from urllib.parse import quote
from concurrent.futures import ThreadPoolExecutor, as_completed

import pandas as pd
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import streamlit as st

st.set_page_config(page_title="기업 스크리너", layout="wide")

st.markdown("""
<style>
[data-testid="stCheckbox"] {
    transform: scale(1.8);
    transform-origin: left center;
}

/* "전체 선택/해제" 체크박스는 위 확대 규칙에서 제외하고, 표 글자 크기와 비슷하게 작게 */
.st-key-select_all_container [data-testid="stCheckbox"] {
    transform: none;
}
.st-key-select_all_container [data-testid="stCheckbox"] label p {
    font-size: 14px !important;
}

/* 메인 타이틀 - 여백/자간 정리 */
h1 {
    letter-spacing: -0.02em;
    padding-top: 0.2rem !important;
}

/* 필터 조건 카드: 테두리를 따라 확실하게 보이는 그림자로 입체감 부여
   (key="filter_box"로 생성되는 .st-key-filter_box 클래스 사용 - 공식 문서에서 권장하는 방식) */
.st-key-filter_box {
    border-radius: 16px !important;
    box-shadow: 0 4px 14px rgba(30, 41, 59, 0.16) !important;
    background: #FFFFFF !important;
}

/* 안내 배너(파란 정보 박스)도 같은 톤으로 살짝 떠 보이게 */
[data-testid="stAlertContainer"] {
    border-radius: 12px !important;
    box-shadow:
        0 1px 2px rgba(30, 41, 59, 0.05),
        0 6px 16px rgba(30, 41, 59, 0.08) !important;
}

/* 각 입력란 위의 라벨 글자: 조금 더 크고 굵게 */
[data-testid="stWidgetLabel"] p {
    font-size: 16px !important;
    font-weight: 700 !important;
}

/* 버튼: 모서리를 부드럽게, 호버 시 미세한 그림자 */
.stButton > button, .stDownloadButton > button {
    border-radius: 8px !important;
    transition: box-shadow 0.15s ease, transform 0.05s ease;
}
.stButton > button:hover, .stDownloadButton > button:hover {
    box-shadow: 0 3px 10px rgba(184, 134, 63, 0.28);
}

/* 결과 요약(성공) 배너의 강조색을 포인트 컬러 톤으로 */
[data-testid="stAlertContentSuccess"] {
    color: #1E3A5F;
}
</style>
""", unsafe_allow_html=True)


class RateLimiter:
    """OpenDART가 초당 요청 수를 제한할 가능성에 대비해, 전체 스레드가 공유하는
    최소 호출 간격을 강제하는 간단한 속도 제한기."""

    def __init__(self, min_interval_sec: float):
        self.min_interval = min_interval_sec
        self.lock = threading.Lock()
        self.last_call = 0.0

    def wait(self):
        with self.lock:
            now = time.monotonic()
            elapsed = now - self.last_call
            if elapsed < self.min_interval:
                time.sleep(self.min_interval - elapsed)
            self.last_call = time.monotonic()


RATE_LIMITER = RateLimiter(min_interval_sec=1 / 15)


def make_session(total=8, backoff_factor=1.5):
    """일시적인 연결 끊김에 대비해 재시도 로직이 포함된 세션 생성."""
    s = requests.Session()
    s.headers.update({
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
            "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
        ),
        "Connection": "close",
    })
    retry = Retry(
        total=total,
        backoff_factor=backoff_factor,
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=["GET"],
    )
    adapter = HTTPAdapter(max_retries=retry, pool_connections=50, pool_maxsize=50)
    s.mount("https://", adapter)
    s.mount("http://", adapter)
    return s


SIDO_MAP = {
    '서울특별시': '서울', '부산광역시': '부산', '대구광역시': '대구', '인천광역시': '인천',
    '광주광역시': '광주', '대전광역시': '대전', '울산광역시': '울산', '세종특별자치시': '세종',
    '경기도': '경기', '강원도': '강원', '강원특별자치도': '강원', '충청북도': '충북', '충청남도': '충남',
    '전라북도': '전북', '전북특별자치도': '전북', '전라남도': '전남', '경상북도': '경북',
    '경상남도': '경남', '제주특별자치도': '제주'
}


def simplify_address(addr: str) -> str:
    if not addr or not isinstance(addr, str):
        return ''
    addr = addr.strip()
    tokens = addr.split()
    if not tokens:
        return addr
    sido = SIDO_MAP.get(tokens[0], tokens[0])
    gu = tokens[1] if len(tokens) > 1 else ''
    if gu and re.match(r'.+(구|시|군)$', gu):
        return f"{sido} {gu}"
    return sido


def _normalize_company_name(name):
    """법인 표기((주), 주식회사, (유), 유한회사)만 떼어낸 순수 회사명.
    검색 API 질의어나 브랜드명 매칭 등, 정식 법인명이 오히려 방해되는 경우에 사용."""
    if not isinstance(name, str):
        return ''
    name = name.strip()
    name = re.sub(r'^\(주\)|\(주\)$|^㈜|㈜$', '', name).strip()
    name = re.sub(r'^주식회사\s*|\s*주식회사$', '', name).strip()
    name = re.sub(r'^\(유\)|\(유\)$', '', name).strip()
    name = re.sub(r'^유한회사\s*|\s*유한회사$', '', name).strip()
    return name


def _download_bytes(session, url, timeout=(20, 60), attempts=3, max_total_seconds=120):
    """용량이 큰 파일도 안정적으로 받기 위해 스트리밍 + 재시도로 다운로드.
    응답이 아주 느리게(찔끔찔끔) 들어오면 개별 read는 타임아웃에 안 걸려도
    전체적으로 계속 늘어질 수 있어, 전체 소요 시간에도 상한선을 둔다."""
    last_err = None
    for attempt in range(attempts):
        start = time.monotonic()
        try:
            with session.get(url, timeout=timeout, stream=True,
                              headers={"Accept-Encoding": "identity"}) as resp:
                resp.raise_for_status()
                chunks = []
                for chunk in resp.iter_content(chunk_size=64 * 1024):
                    if chunk:
                        chunks.append(chunk)
                    if time.monotonic() - start > max_total_seconds:
                        raise TimeoutError(f"전체 다운로드가 {max_total_seconds}초를 넘어 중단함 (응답이 너무 느림)")
                return b"".join(chunks)
        except (requests.exceptions.ConnectionError, TimeoutError) as e:
            last_err = e
            time.sleep(1.5 * (attempt + 1))
    raise RuntimeError(f"다운로드 실패: {last_err}")


DAEBUNRYU_RANGES = [
    ('농업, 임업 및 어업', 1, 3),
    ('광업', 5, 8),
    ('제조업', 10, 34),
    ('전기, 가스, 증기 및 공기 조절 공급업', 35, 35),
    ('수도, 하수 및 폐기물 처리, 원료 재생업', 36, 39),
    ('건설업', 41, 42),
    ('도매 및 소매업', 45, 47),
    ('운수 및 창고업', 49, 52),
    ('숙박 및 음식점업', 55, 56),
    ('정보통신업', 58, 63),
    ('금융 및 보험업', 64, 66),
    ('부동산업', 68, 68),
    ('전문, 과학 및 기술 서비스업', 69, 73),
    ('사업시설 관리, 사업 지원 및 임대 서비스업', 74, 76),
    ('공공 행정, 국방 및 사회보장 행정', 84, 84),
    ('교육 서비스업', 85, 85),
    ('보건업 및 사회복지 서비스업', 86, 87),
    ('예술, 스포츠 및 여가관련 서비스업', 90, 91),
    ('협회 및 단체, 수리 및 기타 개인 서비스업', 94, 96),
    ('가구 내 고용활동 및 달리 분류되지 않은 자가소비 생산활동', 97, 98),
    ('국제 및 외국기관', 99, 99),
]


def _code_to_daebunryu(code2):
    try:
        n = int(code2)
    except Exception:
        return '미분류'
    for name, s, e in DAEBUNRYU_RANGES:
        if s <= n <= e:
            return name
    return '미분류'


@st.cache_data(show_spinner=False, ttl=60 * 60 * 24 * 7)
def load_industry_hierarchy(industry_names):
    """업종명(정밀) -> {대분류, 중분류, 소분류, 세분류} 매핑.
    KSIC 코드표를 받아 역매핑한 뒤, 코드 자릿수를 줄여가며 각 단계 이름을 구한다.
    표기가 살짝 달라 매칭 안 되는 업종명은 모든 단계가 '미분류'로 묶여, 검색 자체는 계속 가능하다."""
    try:
        s = make_session(total=3, backoff_factor=1.0)
        resp = s.get("https://raw.githubusercontent.com/FinanceData/KSIC/master/KSIC_09.csv.gz",
                      timeout=(20, 30))
        resp.raise_for_status()
        ksic = pd.read_csv(io.BytesIO(resp.content), compression='gzip', dtype='str')
        name_to_code = {}
        code_to_name = {}
        for code, name in zip(ksic['Industy_code'], ksic['Industy_name']):
            name = name.strip()
            code_to_name[code] = name
            if name not in name_to_code or len(code) > len(name_to_code[name]):
                name_to_code[name] = code
    except Exception:
        name_to_code, code_to_name = {}, {}

    mapping = {}
    for nm in industry_names:
        code = name_to_code.get(nm)
        if not code:
            mapping[nm] = {'대분류': '미분류', '중분류': '미분류', '소분류': '미분류', '세분류': '미분류'}
            continue
        mapping[nm] = {
            '대분류': _code_to_daebunryu(code[:2]),
            '중분류': code_to_name.get(code[:2], '미분류'),
            '소분류': code_to_name.get(code[:3], '미분류'),
            '세분류': code_to_name.get(code[:4], '미분류'),
        }
    return mapping


@st.cache_data(show_spinner=False, ttl=60 * 60 * 24)
def load_dart_full_registry():
    """DART 기업개황(업종별) 화면의 '엑셀일괄' 다운로드를 그대로 받아온다.
    상장/비상장 포함 DART 등록법인 전체(약 11만개+)에 실제 업종명이 붙어 있어,
    KRX 상장법인목록보다 훨씬 정확하고 넓은 범위를 커버한다. API 키 불필요."""
    session = make_session(total=3, backoff_factor=1.5)
    # 서버가 세션(쿠키) 기반으로 동작할 수 있어, 검색 화면을 먼저 방문한 뒤 같은 세션으로 다운로드한다
    try:
        session.get("https://dart.fss.or.kr/dsae001/main.do", timeout=(20, 20))
    except Exception:
        pass

    content = _download_bytes(
        session, "https://dart.fss.or.kr/dsae001/downloadExcel.do", timeout=(20, 90),
        attempts=2, max_total_seconds=90,
    )
    df = pd.read_excel(io.BytesIO(content))

    df['종목코드'] = df['종목코드'].astype(str).str.strip()
    df.loc[df['종목코드'] == '', '종목코드'] = None
    df['회사명'] = df['회사이름'].astype(str).str.strip()
    df['본사_위치'] = df['주소'].apply(simplify_address)
    df['법인구분'] = df['법인구분'].replace({
        '유가증권시장': '코스피',
        '코스닥시장': '코스닥',
        '코넥스시장': '코넥스',
        '기타법인': '비상장',
    })

    def _year(x):
        try:
            return int(str(x)[:4])
        except Exception:
            return None
    df['설립연도'] = df['설립일'].apply(_year)

    # 상장사는 KRX KIND의 '주요제품' 텍스트를 대표상품/브랜드로 병합 (비상장은 자동 소스가 없음)
    try:
        kind_df = load_kind_key_products()
        df = df.merge(kind_df, on='종목코드', how='left')
    except Exception:
        df['대표상품_브랜드'] = None

    # 자동으로는 못 잡거나 부정확한, 잘 알려진 회사들은 실제 브랜드명으로 덮어씀
    # (법인 표기(주/유 등)만 떼어내고, 그 나머지가 완전히 일치할 때만 적용 -> 짧은 이름의 오매칭 방지)
    _normalized_names = df['회사명'].apply(_normalize_company_name)
    _override_series = _normalized_names.map(BRAND_OVERRIDES)
    df['대표상품_브랜드'] = _override_series.combine_first(df['대표상품_브랜드'])

    return df


FTC_REGIONS = [
    '서울특별시', '부산광역시', '대구광역시', '인천광역시', '광주광역시', '대전광역시',
    '울산광역시', '세종특별자치시', '경기도', '강원특별자치도', '충청북도', '충청남도',
    '전북특별자치도', '전라남도', '경상북도', '경상남도', '제주특별자치도',
]


@st.cache_data(show_spinner=False, ttl=60 * 60 * 24)
def load_ftc_bizcomm_registry():
    """공정거래위원회 '통신판매사업자' 전체 명단을 지역별로 나눠 받아서 하나로 합친다.
    사업자등록번호(숫자만) 집합을 돌려주며, 이 안에 있으면 통신판매업 등록된 것으로 본다.
    지역들을 병렬로 받고, 지역 하나가 실패해도(느려도) 빨리 포기하고 나머지는 계속 진행한다."""
    url = "https://www.ftc.go.kr/www/downloadBizComm.do"

    def _fetch_region(region):
        session = make_session(total=1, backoff_factor=0.5)
        filename = f"통신판매사업자_ALL_{region} 전체.csv"
        params = {"atchFileUrl": "dataopen", "atchFileNm": filename}
        try:
            resp = session.get(url, params=params, timeout=(10, 20))
            resp.raise_for_status()
            content = resp.content
            for enc in ('cp949', 'euc-kr', 'utf-8-sig', 'utf-8'):
                try:
                    df = pd.read_csv(io.BytesIO(content), encoding=enc)
                    if '사업자등록번호' in df.columns:
                        return region, df['사업자등록번호'].astype(str).str.replace(r'\D', '', regex=True)
                except Exception:
                    continue
            return region, None
        except Exception:
            return region, None

    brnos = set()
    ok_regions = []
    with ThreadPoolExecutor(max_workers=10) as executor:
        futures = [executor.submit(_fetch_region, r) for r in FTC_REGIONS]
        for fut in as_completed(futures):
            region, digits = fut.result()
            if digits is not None:
                brnos.update(d for d in digits if len(d) == 10)
                ok_regions.append(region)
    if not ok_regions:
        raise RuntimeError("통신판매업 등록 자료를 하나도 못 받아왔습니다 (지역 17개 전부 실패).")
    return brnos, ok_regions


@st.cache_data(show_spinner=False, ttl=60 * 60 * 24)
def load_kind_key_products():
    """KRX KIND 상장법인목록에서 '주요제품' 텍스트만 가져온다 (상장사 대상, 브랜드명이 섞여 있는 경우가 많음).
    실패해도 전체 로딩을 막지 않도록 재시도/타임아웃을 짧게 잡음."""
    kind_url = "https://kind.krx.co.kr/corpgeneral/corpList.do?method=download&searchType=13"
    s = make_session(total=1, backoff_factor=0.5)
    resp = s.get(kind_url, timeout=(10, 15))
    resp.raise_for_status()
    resp.encoding = 'euc-kr'
    kdf = pd.read_html(io.StringIO(resp.text), header=0)[0]
    kdf['종목코드'] = kdf['종목코드'].astype(str).str.zfill(6)
    kdf = kdf.rename(columns={'주요제품': '대표상품_브랜드'})
    return kdf[['종목코드', '대표상품_브랜드']]


# 자동으로 못 잡거나 부정확한 유명 회사들의 실제 브랜드명 (회사명: '브랜드1, 브랜드2, ...')
# 회사명은 DART 등록명과 정확히 일치해야 적용됨 (상장/비상장 상관없이 전체 적용)
BRAND_OVERRIDES = {
    'LG생활건강': '후, 오휘, 숨, 빌리프, 엘라스틴, 페리오, 샤프란, 테크',
    '애경산업': "Age 20's, 루나, 트리오, 2080치약, 순샘, 케라시스",
    '유한킴벌리': '하기스, 크리넥스, 좋은느낌, 스카트',
    '한국피앤지판매': '페브리즈, 오랄비, 팬틴, 헤드앤숄더, 다우니, 질레트',
    '유니레버코리아': '도브, 바세린, 폰즈, 럭스',
    '로레알코리아': '랑콤, 키엘, 로레알파리, 메이블린뉴욕, 라로슈포제',
    '유한크로락스': '유한락스, 유한젠, 유한 펑크린',
    '옥시레킷벤키저': '옥시크린, 물먹는하마, 데톨, 피니시',
    '한국인삼공사': '정관장, 홍삼정, 에브리타임, 화애락',
    '종근당건강': '락토핏, 아임비타, 아이클리어',
    'CJ웰케어': '바이오코어(유산균), BYO',
    'hy': '헬리코박터 프로젝트 윌, 메치니코프, 하루야채',
    'SK매직': '정수기, 공기청정기, 전기레인지, 매직쿡',
    '쿠쿠전자': '쿠쿠 IH압력밥솥, 전기보온밥솥',
    '청호나이스': '얼음정수기, 공기청정기, 비데, 연수기',
    '다이슨코리아': '무선청소기, 헤어드라이어(에어랩), 공기청정기',
    '필립스코리아': '에어프라이어, 전동칫솔(소닉케어), 헤어케어 가전',
    '세라젬': '팔콘, 퀀텀, 파라오로보 (안마의자/안마베드)',
    '바디프랜드': '팬텀로보, 파라오, 로보워킹 (안마의자)',
    '오텍캐리어': '캐리어에어컨',
    '콜러노비타': '노비타, 필다임 (비데, 가습기, 전기밥솥)',
    '해피콜': '다이아몬드 프라이팬, IH진공냄비',

    # 화장품 기업 및 운영 브랜드 (사용자 제공 목록, 47개사)
    '달바글로벌': "d'Alba (달바)",
    '아모레퍼시픽': '설화수, 라네즈, 헤라, 아이오페, 마몽드, 에뛰드, 이니스프리, 한율, 에스트라 등',
    '에이블씨엔씨': "미샤(MISSHA), 어퓨(A'pieu), 초공진, 스틸라, 셀라피",
    '에이피알': '메디큐브(medicube), 에이프릴스킨(APRILSKIN), 포맨트(FORMENT), 글램디바이오',
    '잇츠한불': "잇츠스킨(It's SKIN), ICS, 체이싱래빗",
    '한불화장품': "잇츠스킨(It's SKIN), ICS, 체이싱래빗",
    '토니모리': '토니모리(TONYMOLY), 튜티',
    '네오팜': '아토팜(ATOPALM), 리얼베리어, 제로이드, 더마비',
    '삐아': '삐아(BBIA), 어바웃톤, 이글립스, 에디트B',
    '아로마티카': '아로마티카(AROMATICA)',
    '네이처리퍼블릭': '네이처리퍼블릭(NATURE REPUBLIC)',
    '논픽션': '논픽션(NONFICTION)',
    '더스킨팩토리': '쿵스파파, 쿤달(KUNDAL)',
    '더파운더즈': '아누아(ANUA)',
    '더페이스샵': '더페이스샵',
    '데이지크': '데이지크(dasique)',
    '두리화장품': '댕기머리',
    '리만코리아': '인셀덤(INCELLDERM), 보타랩, 리프토라인',
    '믹스앤매치': '믹순(mixsoon)',
    '비나우': '넘버즈인(numbuzin), 퓌(fwee)',
    '비앤에이치코스메틱': '지베르니, 아크웰',
    '비에이치랩': '모다모다, 블랙모체 등 헤어케어 브랜드',
    '서린컴퍼니': '라운드랩(ROUND LAB - 독도토너 등)',
    '세화피앤씨': '모레모(moremo)',
    '스킨79': '스킨79',
    '스킨이데아': '메디필(MEDI-PEEL)',
    '시드물': '시드물(SIDMOOL)',
    '쏘내추럴': "쏘내추럴(SO'NATURAL)",
    '씨엠에스랩': '셀퓨전씨(Cell Fusion C)',
    '에프앤코': '바닐라코(BANILA CO)',
    '엔프라니': '엔프라니, 홀리카홀리카(Holika Holika)',
    '이즈앤트리': '이즈앤트리(ISNTREE)',
    '자연인': '아이소이(isoi)',
    '정샘물뷰티': '정샘물(JUNG SAEM MOOL)',
    '제이숲': '제이숲(JSOOP)',
    '참존': '참존(CHARMZONE)',
    '초초스팩토리': '조성아22, 16브랜드',
    '코스알엑스': '코스알엑스(COSRX)',
    '코스모코스': '꽃을든남자, 다나한, 에프엠식스',
    '토리든': '토리든(Torriden)',
    '투쿨포스쿨': '투쿨포스쿨(too cool for school)',
    '티에스트릴리온': 'TS샴푸',
    '페렌벨': '썸바이미(SOME BY MI)',
    '편강한방피부과학연구소': '편강율(Pyunkang Yul)',
    '포컴퍼니': '아비브(Abib)',
    '해브앤비': '닥터자르트(Dr.Jart+)',
    '휴메이저': '닥터포헤어 등 헤어케어 브랜드',

    # 화장품 OEM/ODM 전문 제조기업 (브랜드가 아니라 위탁생산 전문 영역)
    '코스맥스': 'OEM/ODM 전문 기업 (타사 화장품 위탁 생산)',
    '한국콜마': 'OEM/ODM 전문 기업 (타사 화장품 위탁 생산)',
    '한국화장품제조': 'OEM/ODM 전문 기업 (타사 화장품 위탁 생산)',
    '씨앤씨인터내셔널': 'OEM/ODM 전문 기업 (색조 화장품 위탁 생산)',
    '코스메카코리아': 'OEM/ODM 전문 기업 (기초 및 색조 위탁 생산)',
    '잉글우드랩': 'OEM/ODM 전문 기업 (기초 및 색조 위탁 생산)',
    '코스비전': 'OEM/ODM 전문 기업 (아모레퍼시픽 그룹 전속 제조)',
    '라파스': 'OEM/ODM 전문 기업 (마이크로니들 패치 전문)',
    '인터코스': 'OEM/ODM 전문 기업 (글로벌 이탈리아계 ODM)',
    '인터코스코리아': 'OEM/ODM 전문 기업 (글로벌 이탈리아계 ODM)',
    '씨앤에프': 'OEM/ODM 전문 기업 (마스크팩 및 기초 전문 ODM)',
    '메가코스': 'OEM/ODM 전문 기업 (토니모리 계열 제조법인)',
    '나우코스': 'OEM/ODM 전문 기업들',
    '나투젠': 'OEM/ODM 전문 기업들',
    '씨앤텍': 'OEM/ODM 전문 기업들',
    '셀랩': 'OEM/ODM 전문 기업들',
}


@st.cache_data(show_spinner=False, ttl=60 * 60 * 24)
def load_corp_map(api_key: str):
    """OpenDART corpCode.xml -> {종목코드: corp_code} 매핑. 매출 조회 시에만 필요."""
    import zipfile
    import xml.etree.ElementTree as ET

    url = f"https://opendart.fss.or.kr/api/corpCode.xml?crtfc_key={api_key}"
    session = make_session(total=3, backoff_factor=1.5)
    content = _download_bytes(session, url, timeout=(20, 60))

    try:
        zf = zipfile.ZipFile(io.BytesIO(content))
    except zipfile.BadZipFile:
        # zip이 아니면 보통 OpenDART가 보낸 에러 메시지(JSON/텍스트)일 가능성이 높음 - 그대로 보여줌
        try:
            preview = content.decode('utf-8', errors='replace')[:500]
        except Exception:
            preview = str(content[:200])
        raise RuntimeError(
            f"OpenDART가 정상 파일 대신 이런 응답을 보냈습니다 (API 키를 확인해보세요): {preview}"
        )
    xml_bytes = zf.read(zf.namelist()[0])
    root = ET.fromstring(xml_bytes)
    corp_map = {}
    for child in root.findall('list'):
        stock_code = child.findtext('stock_code', default='').strip()
        corp_code = child.findtext('corp_code', default='').strip()
        if stock_code:
            corp_map[stock_code] = corp_code
    return corp_map


POSITIVE_WORDS = [
    '급등', '호실적', '흑자', '수주', '성장', '역대', '최대', '호평', '선정', '1위',
    '개선', '반등', '상승', '투자유치', '신기록', '돌파', '확대', '진출', '체결', '수출',
]
NEGATIVE_WORDS = [
    '급락', '적자', '부진', '소송', '리콜', '논란', '제재', '하락', '악재', '위기',
    '감소', '철수', '중단', '파산', '구조조정', '횡령', '배임', '과징금', '불매', '피소',
]


ECONOMIC_PRESS_DOMAINS = {
    "매일경제": "mk.co.kr",
    "한국경제": "hankyung.com",
    "서울경제": "sedaily.com",
    "머니투데이": "mt.co.kr",
    "이데일리": "edaily.co.kr",
    "파이낸셜뉴스": "fnnews.com",
    "헤럴드경제": "heraldcorp.com",
    "아시아경제": "asiae.co.kr",
    "아주경제": "ajunews.com",
    "이투데이": "etoday.co.kr",
}



def get_naver_news_analysis(session, client_id, client_secret, company_name, months=6, max_titles_shown=5):
    """NAVER API HUB의 뉴스 검색 API로 최근 기사를 가져와서
    1) 최근 제목 몇 개 (참고용, 사람이 직접 판단)
    2) 최근 N개월 이내 기사 제목에 긍정/부정 단어가 몇 개 나오는지 세는 단순 방식의 5단계 분류
    3) 그 안에 주요 경제지(매일경제/한국경제/서울경제 등 10개) 도메인 기사가 있는지 확인
    을 함께 반환한다. 정교한 감성분석이 아니라 참고용 신호일 뿐임."""
    url = "https://naverapihub.apigw.ntruss.com/search/v1/news"
    headers = {
        "X-NCP-APIGW-API-KEY-ID": client_id,
        "X-NCP-APIGW-API-KEY": client_secret,
    }
    params = {"query": company_name, "display": 30, "start": 1, "sort": "date", "format": "json"}
    try:
        RATE_LIMITER.wait()
        r = session.get(url, headers=headers, params=params, timeout=(15, 15))
        if r.status_code != 200:
            err = f"NAVER API 오류: {r.status_code}"
            return err, err, err
        data = r.json()
        items = data.get('items', [])
        if not items:
            return '검색 결과 없음', '검색 결과 없음', '검색 결과 없음'

        titles_all = [re.sub('<[^<]+?>', '', it.get('title', '')) for it in items]
        titles_preview = ' / '.join(titles_all[:max_titles_shown])

        cutoff = pd.Timestamp.now(tz='Asia/Seoul') - pd.DateOffset(months=months)
        recent_items = []
        for it in items:
            try:
                dt = pd.to_datetime(it.get('pubDate'))
                if dt.tzinfo is None:
                    dt = dt.tz_localize('Asia/Seoul')
                if dt >= cutoff:
                    recent_items.append(it)
            except Exception:
                continue

        # 경제지 보도 확인 - 원문 링크(originallink) 도메인 기준
        press_hits = []
        for press_name, domain in ECONOMIC_PRESS_DOMAINS.items():
            for it in recent_items:
                link = it.get('originallink') or it.get('link', '')
                if domain in link:
                    press_hits.append(press_name)
                    break
        press_result = ", ".join(press_hits) + " 보도 확인" if press_hits else "확인 안됨"

        if not recent_items:
            return titles_preview, f'최근 {months}개월 내 기사 없음', press_result

        recent_titles = [re.sub('<[^<]+?>', '', it.get('title', '')) for it in recent_items]
        pos = neg = 0
        for title in recent_titles:
            for w in POSITIVE_WORDS:
                if w in title:
                    pos += 1
            for w in NEGATIVE_WORDS:
                if w in title:
                    neg += 1
        score = pos - neg
        if score >= 3:
            label = '매우 긍정'
        elif score >= 1:
            label = '긍정'
        elif score == 0:
            label = '보통'
        elif score >= -2:
            label = '부정'
        else:
            label = '매우 부정'
        return titles_preview, label, press_result
    except Exception as e:
        err = f"연결 오류: {e}"
        return err, err, err


@st.dialog("🔎 검색 처리 중", width="large")
def _search_processing_dialog(candidates, fetch_revenue, api_key, bsns_year, max_workers,
                               min_rev, max_rev, min_op, max_op, min_ni, max_ni, top_n,
                               fetch_titles, naver_client_id, naver_client_secret):
    """검색 버튼을 누른 뒤의 모든 처리(매출조회/최근뉴스 포함)를 화면 가운데 팝업 안에서 진행하고,
    끝나면 결과를 세션에 저장한 뒤 다시 그림 (선택 옵션과 무관하게 항상 이 팝업을 거침)."""
    st.write(f"1차 필터 후 후보 기업 수: **{len(candidates)}**")

    final = candidates.copy()
    final['매출액(억원)'] = None
    final['영업이익(억원)'] = None
    final['당기순이익(억원)'] = None
    final['실패사유'] = None
    no_data_df = pd.DataFrame()

    if fetch_revenue:
        st.write("**매출 / 영업이익 / 당기순이익 조회**" + (" (1/2단계)" if fetch_titles else ""))
        try:
            corp_map = load_corp_map(api_key)
        except Exception as e:
            st.error(f"OpenDART corp_code 매핑을 못 가져왔습니다: {e}")
            st.session_state["still_running"] = False
            st.stop()  # 팝업은 열린 채로 두어 에러 메시지를 보여주고, 사용자가 직접 닫게 함

        listed = final[final['종목코드'].notna()].copy()
        unlisted_count = len(final) - len(listed)
        if unlisted_count > 0:
            st.caption(f"비상장 {unlisted_count}개사는 매출 조회 대상이 아니라 매출 없이 표시됩니다.")

        session = make_session(total=2, backoff_factor=0.5)
        progress = st.progress(0.0, text="매출/영업이익/순이익 조회 중...")
        results = {}
        executor = ThreadPoolExecutor(max_workers=int(max_workers))
        try:
            future_map = {}
            for idx, row in listed.iterrows():
                corp_code = corp_map.get(row['종목코드'])
                if not corp_code:
                    continue
                fut = executor.submit(get_financials, session, api_key, corp_code, bsns_year, "11011")
                future_map[fut] = idx
            done = 0
            for fut in as_completed(future_map):
                idx = future_map[fut]
                results[idx] = fut.result()
                done += 1
                if len(future_map):
                    progress.progress(done / len(future_map),
                                       text=f"매출/영업이익/순이익 조회 중... ({done}/{len(future_map)})")
        finally:
            executor.shutdown(wait=False, cancel_futures=True)

        for idx, (revenue, op_profit, net_income, reason) in results.items():
            final.at[idx, '매출액(억원)'] = (revenue / 1e8) if revenue is not None else None
            final.at[idx, '영업이익(억원)'] = (op_profit / 1e8) if op_profit is not None else None
            final.at[idx, '당기순이익(억원)'] = (net_income / 1e8) if net_income is not None else None
            final.at[idx, '실패사유'] = reason

        no_data_df = final[final['종목코드'].notna() & final['매출액(억원)'].isna()][
            ['회사명', '종목코드', '실패사유']
        ].copy()

        for col, mn, mx in [('매출액(억원)', min_rev, max_rev),
                             ('영업이익(억원)', min_op, max_op),
                             ('당기순이익(억원)', min_ni, max_ni)]:
            final[col] = pd.to_numeric(final[col], errors='coerce')
            final = final[final[col].isna() | ((final[col] >= mn) & (final[col] <= mx))]

    final = final.sort_values(
        '매출액(억원)', ascending=False, na_position='last'
    ).head(int(top_n)).reset_index(drop=True)
    for col in ['매출액(억원)', '영업이익(억원)', '당기순이익(억원)']:
        final[col] = pd.to_numeric(final[col], errors='coerce').round(0).astype('Int64')

    output_df = final.rename(columns={
        '업종명': '업종', '홈페이지': '홈페이지 주소', '본사_위치': '본사 위치',
    })
    output_df['관련기사'] = output_df['회사명'].apply(
        lambda name: f"https://search.naver.com/search.naver?where=news&query={quote(_normalize_company_name(str(name)) or str(name))}"
    )

    def bizno_url(brno):
        digits = re.sub(r'\D', '', str(brno)) if pd.notna(brno) else ''
        return f"https://bizno.net/article/{digits}" if len(digits) == 10 else None

    output_df['기업정보(bizno)'] = final['사업자등록번호'].apply(bizno_url)

    def naver_finance_url(stock_code):
        if pd.isna(stock_code) or not str(stock_code).strip():
            return None
        return f"https://finance.naver.com/item/main.naver?code={str(stock_code).strip()}"

    output_df['증권'] = final['종목코드'].apply(naver_finance_url)

    def ftc_biz_url(row):
        brno = row['사업자등록번호']
        digits = re.sub(r'\D', '', str(brno)) if pd.notna(brno) else ''
        if len(digits) != 10 or not row.get('통신판매업_등록', False):
            return None
        return f"http://www.ftc.go.kr/bizCommPop.do?wrkr_no={digits}"

    output_df['통신판매업조회'] = final.apply(ftc_biz_url, axis=1)

    output_df['대분류'] = final['대분류']
    output_df['대표상품/브랜드'] = final['대표상품_브랜드']

    _base_cols = ['회사명', '대표상품/브랜드', '관련기사', '기업정보(bizno)', '증권', '통신판매업조회', '홈페이지 주소', '대분류', '업종',
                  '법인구분', '대표자명', '설립연도', '본사 위치']
    if fetch_revenue:
        _base_cols = ['회사명', '대표상품/브랜드', '관련기사', '기업정보(bizno)', '증권', '통신판매업조회', '홈페이지 주소', '대분류', '업종',
                      '법인구분', '대표자명', '매출액(억원)', '영업이익(억원)', '당기순이익(억원)',
                      '설립연도', '본사 위치']
    output_df = output_df[_base_cols]

    if fetch_titles:
        if not naver_client_id or not naver_client_secret:
            st.warning("최근 뉴스를 켜셨다면 NAVER API HUB Client ID/Secret을 입력해주세요. (이 항목은 건너뜁니다)")
        else:
            st.write("**최근 뉴스 조회**" + (" (2/2단계)" if fetch_revenue else ""))
            titles_session = make_session(total=2, backoff_factor=0.5)
            progress2 = st.progress(0.0, text="최근 뉴스 조회 중...")
            names = output_df['회사명'].tolist()
            search_names = [_normalize_company_name(nm) or nm for nm in names]
            title_results = [None] * len(names)
            sentiment_results = [None] * len(names)
            press_results = [None] * len(names)
            executor2 = ThreadPoolExecutor(max_workers=10)
            try:
                future_map = {
                    executor2.submit(
                        get_naver_news_analysis, titles_session, naver_client_id, naver_client_secret, sn
                    ): i
                    for i, sn in enumerate(search_names)
                }
                done = 0
                for fut in as_completed(future_map):
                    idx = future_map[fut]
                    titles_preview, sentiment_label, press_result = fut.result()
                    title_results[idx] = titles_preview
                    sentiment_results[idx] = sentiment_label
                    press_results[idx] = press_result
                    done += 1
                    progress2.progress(done / len(names), text=f"최근 뉴스 조회 중... ({done}/{len(names)})")
            finally:
                executor2.shutdown(wait=False, cancel_futures=True)

            output_df['뉴스여론'] = sentiment_results
            output_df['경제지 보도(10개 매체)'] = press_results

    _desired_order = [
        '회사명', '대표상품/브랜드', '기업정보(bizno)', '증권', '통신판매업조회', '홈페이지 주소', '관련기사',
        '뉴스여론', '경제지 보도(10개 매체)',
        '대분류', '업종', '법인구분',
        '매출액(억원)', '영업이익(억원)', '당기순이익(억원)', '대표자명', '설립연도', '본사 위치',
    ]
    _final_order = [c for c in _desired_order if c in output_df.columns]
    _final_order += [c for c in output_df.columns if c not in _final_order]
    output_df = output_df[_final_order]

    st.session_state["last_output_df"] = output_df
    st.session_state["_no_data_df"] = no_data_df
    st.session_state["_select_all_value"] = False
    st.session_state["_editor_key_counter"] = st.session_state.get("_editor_key_counter", 0) + 1
    st.session_state["still_running"] = False
    st.rerun()


def get_financials(session, api_key, corp_code, bsns_year, reprt_code):
    """fnlttSinglAcnt(주요계정) API - 매출액/영업이익/당기순이익만 가볍게 조회"""
    url = "https://opendart.fss.or.kr/api/fnlttSinglAcnt.json"
    params = {"crtfc_key": api_key, "corp_code": corp_code, "bsns_year": bsns_year, "reprt_code": reprt_code}
    try:
        RATE_LIMITER.wait()
        r = session.get(url, params=params, timeout=(20, 15))
        data = r.json()
        if data.get('status') != '000':
            reason = f"OpenDART 응답: {data.get('status')} - {data.get('message', '')}"
            return None, None, None, reason
        revenue = revenue_fallback = op_profit = net_income = None
        for it in data.get('list', []):
            nm = it.get('account_nm', '')
            amt_str = it.get('thstrm_amount', '').replace(',', '')
            try:
                amt = int(amt_str)
            except ValueError:
                continue
            if nm in ('매출액', '수익(매출액)') and revenue is None:
                revenue = amt
            elif nm == '영업수익' and revenue_fallback is None:
                revenue_fallback = amt
            elif nm == '영업이익' and op_profit is None:
                op_profit = amt
            elif nm in ('당기순이익', '당기순이익(손실)') and net_income is None:
                net_income = amt
        if revenue is None and revenue_fallback is not None:
            revenue = revenue_fallback
        if revenue is None:
            return None, op_profit, net_income, (
                "OpenDART 응답은 정상이지만 '매출액'에 해당하는 계정을 못 찾음"
            )
        return revenue, op_profit, net_income, None
    except Exception as e:
        return None, None, None, f"연결 오류: {e}"


# ---------------- UI ----------------

st.title("📊 조건 기반 기업 스크리너")
st.caption("DART 등록법인 전체(상장+비상장) 기준으로 업종/지역/설립연도 등을 필터링합니다. "
           "매출 조회는 상장사에 한해 선택적으로 가능합니다.")

with st.spinner("DART 전체법인 목록 불러오는 중... (최초 1회, 파일이 커서 1분 정도 걸릴 수 있습니다)"):
    try:
        registry_df = load_dart_full_registry()
        load_error = None
    except Exception as e:
        registry_df = None
        load_error = str(e)

if load_error:
    st.error(
        f"DART 전체법인 목록을 못 가져왔습니다: {load_error}\n\n"
        "잠시 후 새로고침해서 다시 시도해보세요."
    )
    st.stop()

industry_options = sorted(registry_df['업종명'].dropna().astype(str).str.strip().unique().tolist())
industry_options = [x for x in industry_options if x]
corp_type_options = sorted(registry_df['법인구분'].dropna().astype(str).unique().tolist())

with st.spinner("업종 계층(대/중/소/세분류) 매핑 중... (최초 1회)"):
    hierarchy_map = load_industry_hierarchy(tuple(industry_options))
registry_df['대분류'] = registry_df['업종명'].astype(str).str.strip().map(
    lambda x: hierarchy_map.get(x, {}).get('대분류', '미분류')
)
registry_df['중분류'] = registry_df['업종명'].astype(str).str.strip().map(
    lambda x: hierarchy_map.get(x, {}).get('중분류', '미분류')
)
registry_df['소분류'] = registry_df['업종명'].astype(str).str.strip().map(
    lambda x: hierarchy_map.get(x, {}).get('소분류', '미분류')
)
_order = [nm for nm, _, _ in DAEBUNRYU_RANGES] + ['미분류']
_present = set(v['대분류'] for v in hierarchy_map.values())
daebunryu_options = [x for x in _order if x in _present]

with st.spinner("통신판매업 등록 명단 불러오는 중... (최초 1회, 지역 17개라 1~2분 걸릴 수 있습니다)"):
    try:
        _ftc_brnos, _ftc_ok_regions = load_ftc_bizcomm_registry()
        ftc_load_error = None
    except Exception as e:
        _ftc_brnos = set()
        _ftc_ok_regions = []
        ftc_load_error = str(e)
registry_df['_brno_digits'] = registry_df['사업자등록번호'].astype(str).str.replace(r'\D', '', regex=True)
registry_df['통신판매업_등록'] = registry_df['_brno_digits'].isin(_ftc_brnos)

# 결과표 아래 "재검색" 버튼에서 넘어온 값이 있으면, 위젯이 그려지기 전에 미리 반영
if "pending_industry_search" in st.session_state:
    _picked = st.session_state.pop("pending_industry_search")  # 업종명 리스트

    # 이전 검색에 남아있던 다른 필터들이 AND 조건으로 계속 걸려서 결과가 이상해지는 걸 막기 위해
    # 업종 재검색 시에는 관련 없는 필터를 전부 초기화함
    st.session_state["company_name_search_input"] = ""
    st.session_state["ceo_name_search_input"] = ""
    st.session_state["corp_type_filter_input"] = []
    st.session_state["region_filter_input"] = ""
    st.session_state["min_founding_year_input"] = 0

    # 계단식 필터(대/중/소분류)에도 정확히 반영
    _picked_in_map = [nm for nm in _picked if nm in hierarchy_map]
    st.session_state["daebunryu_select_input"] = sorted(set(
        hierarchy_map[nm]['대분류'] for nm in _picked_in_map
    ))
    st.session_state["jungbunryu_select_input"] = sorted(set(
        hierarchy_map[nm]['중분류'] for nm in _picked_in_map
    ))
    st.session_state["sobunryu_select_input"] = sorted(set(
        hierarchy_map[nm]['소분류'] for nm in _picked_in_map
    ))
    st.session_state["industry_select_input"] = sorted(set(_picked))

    # 정확히 일치하는 업종을 4단계에서 이미 골랐으므로, 키워드 검색창은 비워서 중복/혼선 방지
    st.session_state["industry_keywords_input"] = ""
    st.session_state["enter_pressed_search"] = True

FILTER_DEFAULTS = {
    "daebunryu_select_input": [],
    "jungbunryu_select_input": [],
    "sobunryu_select_input": [],
    "industry_select_input": [],
    "industry_keywords_input": "",
    "company_name_search_input": "",
    "ceo_name_search_input": "",
    "corp_type_filter_input": [],
    "region_filter_input": "",
    "min_founding_year_input": 0,
    "top_n_input": 200,
}

if st.session_state.get("pending_reset_filters", False):
    for k, v in FILTER_DEFAULTS.items():
        st.session_state[k] = v
    st.session_state["pending_reset_filters"] = False

show_results = "last_output_df" in st.session_state

if not show_results:
    st.info("아래의 필터 조건을 입력하고 '검색 실행' 버튼을 누르세요.")

if show_results:
    # 검색 결과가 있으면 필터는 사이드바로 (결과 화면을 넓게 쓰기 위해)
    header_area = st.sidebar
    col_a = col_b = col_c = st.sidebar
else:
    # 처음 켰을 때(검색 전)는 필터를 메인 화면에 3열로 넓게 배치
    header_area = st.container(border=True, key="filter_box")

with header_area:
    if show_results:
        # 검색 후(사이드바)에서는 기존처럼 단순한 스타일 유지
        st.header("🔎 필터 조건")
        if st.button("🔄 필터 초기화", use_container_width=True):
            st.session_state["pending_reset_filters"] = True
            st.session_state.pop("last_output_df", None)
            st.rerun()
    else:
        # 검색 전(메인 3열 화면)에서만 버튼을 크게, 우측 상단에 배치
        _title_col, _reset_col = st.columns([4, 2])
        _title_col.header("🔎 필터 조건")
        _reset_col.write("")  # 헤더 텍스트와 세로 위치를 맞추기 위한 여백
        if _reset_col.button("🔄 필터 초기화", use_container_width=True, type="primary"):
            st.session_state["pending_reset_filters"] = True
            st.session_state.pop("last_output_df", None)
            st.rerun()
    st.caption("업종 대분류→중분류→소분류를 계단식으로 좁히거나, 바로 아래 키워드 검색으로 넓게 찾을 수 있습니다.")

    if not show_results:
        col_a, col_b, col_c = st.columns(3)

    daebunryu_select = col_a.multiselect(
        "1단계: 업종 대분류", options=daebunryu_options, default=[],
        key="daebunryu_select_input",
    )
    _pool = [nm for nm, h in hierarchy_map.items() if not daebunryu_select or h['대분류'] in daebunryu_select]

    jungbunryu_options = sorted(set(hierarchy_map[nm]['중분류'] for nm in _pool))
    jungbunryu_select = col_a.multiselect(
        "2단계: 중분류", options=jungbunryu_options, default=[],
        key="jungbunryu_select_input",
    )
    if jungbunryu_select:
        _pool = [nm for nm in _pool if hierarchy_map[nm]['중분류'] in jungbunryu_select]

    sobunryu_options = sorted(set(hierarchy_map[nm]['소분류'] for nm in _pool))
    sobunryu_select = col_a.multiselect(
        "3단계: 소분류", options=sobunryu_options, default=[],
        key="sobunryu_select_input",
    )
    if sobunryu_select:
        _pool = [nm for nm in _pool if hierarchy_map[nm]['소분류'] in sobunryu_select]

    industry_select = col_a.multiselect(
        "4단계: 업종 선택 (DART 정밀 업종명)",
        options=sorted(set(_pool) | set(st.session_state.get("industry_select_input", []))),
        default=[],
        key="industry_select_input",
    )
    industry_keywords = col_a.text_input(
        "업종 키워드 검색 (콤마로 구분, Enter로 바로 검색)", "",
        key="industry_keywords_input",
        on_change=lambda: st.session_state.update({"enter_pressed_search": True}),
    )
    company_name_search = col_b.text_input(
        "회사명 검색 (콤마로 여러 개 가능, Enter로 바로 검색)", "",
        key="company_name_search_input",
        on_change=lambda: st.session_state.update({"enter_pressed_search": True}),
    )
    ceo_name_search = col_b.text_input(
        "대표자명 검색 (Enter로 바로 검색)", "", key="ceo_name_search_input",
        on_change=lambda: st.session_state.update({"enter_pressed_search": True}),
    )
    corp_type_filter = col_b.multiselect(
        "법인구분", options=corp_type_options, default=[],
        help="코스피/코스닥/코넥스 = 상장사, 비상장 = 외감대상 비상장법인",
        key="corp_type_filter_input",
    )
    region_filter = col_b.text_input("본사 지역 (예: 서울)", "", key="region_filter_input")
    min_founding_year = col_b.number_input(
        "설립연도 (이후 설립된 기업만, 0=필터 없음)", 0, 2100, 0, key="min_founding_year_input"
    )
    if ftc_load_error:
        col_b.caption(f"⚠️ 통신판매업 명단을 못 가져왔습니다: {ftc_load_error}")
    elif len(_ftc_ok_regions) < len(FTC_REGIONS):
        col_b.caption(f"⚠️ 통신판매업 명단 중 {len(_ftc_ok_regions)}/{len(FTC_REGIONS)}개 지역만 불러와졌습니다 (일부 지역 누락 가능).")
    top_n = col_c.number_input("최대 결과 개수", 1, 2000, 200, key="top_n_input")

    col_c.markdown("---")
    _rev_title, _rev_check = col_c.columns([4, 1])
    _rev_title.markdown("**💰 매출조회 (상장사만 가능)**")
    fetch_revenue = _rev_check.checkbox("매출조회", value=False, label_visibility="collapsed")
    col_c.caption("매출 / 영업이익 / 당기순이익이 표시됩니다.")
    api_key = ""
    bsns_year = "2025"
    min_rev = max_rev = min_op = max_op = min_ni = max_ni = None
    max_workers = 10
    if fetch_revenue:
        try:
            _secret_key = st.secrets.get("OPENDART_API_KEY", "")
        except Exception:
            _secret_key = ""
        if _secret_key:
            api_key = _secret_key
            col_c.caption("✅ 저장된 API 키를 사용합니다 (Secrets에 등록됨)")
        else:
            api_key = col_c.text_input("OpenDART API 키", type="password")
        bsns_year = col_c.text_input("조회 사업연도", "2025")
        col_c.markdown("**매출액(억원)**")
        c1, c2 = col_c.columns(2)
        min_rev = c1.number_input("최소", value=0, step=100, key="min_rev")
        max_rev = c2.number_input("최대", value=10000000, step=100, key="max_rev")
        col_c.markdown("**영업이익(억원)**")
        c1, c2 = col_c.columns(2)
        min_op = c1.number_input("최소", value=-10000000, step=100, key="min_op")
        max_op = c2.number_input("최대", value=10000000, step=100, key="max_op")
        col_c.markdown("**당기순이익(억원)**")
        c1, c2 = col_c.columns(2)
        min_ni = c1.number_input("최소", value=-10000000, step=100, key="min_ni")
        max_ni = c2.number_input("최대", value=10000000, step=100, key="max_ni")
        max_workers = col_c.number_input("동시 요청 수", 1, 30, 10)

    col_c.markdown("---")
    _news_title, _news_check = col_c.columns([4, 1])
    _news_title.markdown("**📰 최근뉴스 (참고용)**")
    fetch_titles = _news_check.checkbox("최근뉴스", value=False, label_visibility="collapsed")
    col_c.caption(
        "최근 6개월 네이버에 보도된 뉴스의 키워드를 분석하여, 기사 논조를 5단계 (매우긍정~매우부정)로 표시합니다.\n\n"
        "한경, 매경 등 10개 경제지 매체 보도 여부를 나타냅니다."
    )
    naver_client_id = ""
    naver_client_secret = ""
    if fetch_titles:
        try:
            _secret_id = st.secrets.get("NAVER_CLIENT_ID", "")
            _secret_secret = st.secrets.get("NAVER_CLIENT_SECRET", "")
        except Exception:
            _secret_id, _secret_secret = "", ""
        if _secret_id and _secret_secret:
            naver_client_id, naver_client_secret = _secret_id, _secret_secret
            col_c.caption("✅ 저장된 NAVER API 키를 사용합니다 (Secrets에 등록됨)")
        else:
            naver_client_id = col_c.text_input("NAVER API HUB Client ID")
            naver_client_secret = col_c.text_input("NAVER API HUB Client Secret", type="password")
            col_c.caption("ncloud.com → NAVER API HUB → 뉴스 검색 API 신청 후 발급받을 수 있습니다.")

    run_btn_clicked = st.button("🚀 검색 실행", type="primary", use_container_width=True)
    run_btn = run_btn_clicked or st.session_state.get("enter_pressed_search", False)

COOLDOWN_SEC = 65
if st.session_state.get("still_running", False):
    elapsed = time.time() - st.session_state.get("last_dispatch_ts", 0)
    remaining = COOLDOWN_SEC - elapsed
    if remaining > 0:
        st.warning(f"⏳ 이전 검색이 중간에 중단된 것 같습니다. 약 **{int(remaining)}초** 더 기다려주세요.")
        time.sleep(1)
        st.rerun()
    else:
        st.session_state["still_running"] = False

if run_btn:
    st.session_state["enter_pressed_search"] = False

    if fetch_revenue and not api_key:
        st.warning("매출 조회를 켜셨다면 OpenDART API 키를 입력해주세요.")
        st.stop()

    candidates = registry_df.copy()
    kw_list = [k.strip() for k in industry_keywords.split(',') if k.strip()]
    name_list = [n.strip() for n in company_name_search.split(',') if n.strip()]

    has_any_filter = any([
        name_list, kw_list, industry_select, daebunryu_select, jungbunryu_select, sobunryu_select,
        ceo_name_search.strip(), corp_type_filter, region_filter.strip(), min_founding_year > 0,
    ])
    if not has_any_filter:
        st.warning(
            "필터 조건이 하나도 없어서 검색을 진행하지 않았습니다. "
            "업종/회사명/대표자명/지역 등 최소 하나는 입력해주세요. "
            "(조건 없이 검색하면 전체 11만여 개 회사를 대상으로 돌아 매우 오래 걸립니다)"
        )
        st.stop()

    if name_list:
        pattern = '|'.join(re.escape(n) for n in name_list)
        candidates = candidates[candidates['회사명'].str.contains(pattern, na=False)]
    else:
        if daebunryu_select:
            candidates = candidates[candidates['대분류'].isin(daebunryu_select)]
        if jungbunryu_select:
            candidates = candidates[candidates['중분류'].isin(jungbunryu_select)]
        if sobunryu_select:
            candidates = candidates[candidates['소분류'].isin(sobunryu_select)]
        if industry_select or kw_list:
            mask = pd.Series(False, index=candidates.index)
            if industry_select:
                mask = mask | candidates['업종명'].isin(industry_select)
            if kw_list:
                pattern = '|'.join(kw_list)
                mask = mask | candidates['업종명'].astype(str).str.contains(pattern, na=False)
            candidates = candidates[mask]

    if ceo_name_search.strip():
        candidates = candidates[candidates['대표자명'].astype(str).str.contains(ceo_name_search.strip(), na=False)]

    if corp_type_filter:
        candidates = candidates[candidates['법인구분'].isin(corp_type_filter)]

    if min_founding_year and min_founding_year > 0:
        candidates = candidates[candidates['설립연도'].isna() | (candidates['설립연도'] >= min_founding_year)]

    if region_filter.strip():
        candidates = candidates[candidates['본사_위치'].str.contains(region_filter.strip(), na=False)]

    candidates = candidates.reset_index(drop=True)

    if len(candidates) == 0:
        st.warning("조건에 맞는 후보가 없습니다. 필터를 완화해보세요.")
        st.stop()

    st.session_state["still_running"] = True
    st.session_state["last_dispatch_ts"] = time.time()
    _search_processing_dialog(
        candidates, fetch_revenue, api_key, bsns_year, max_workers,
        min_rev, max_rev, min_op, max_op, min_ni, max_ni, top_n,
        fetch_titles, naver_client_id, naver_client_secret,
    )
    st.stop()  # 팝업이 떠 있는 동안 아래 코드가 먼저 실행되지 않도록 함

if "last_output_df" in st.session_state:
    output_df = st.session_state["last_output_df"]

    st.success(f"최종 {len(output_df)}개사")

    _no_data_df = st.session_state.get("_no_data_df")
    if _no_data_df is not None and not _no_data_df.empty:
        with st.expander(f"⚠️ 매출 데이터를 못 가져온 상장사 {len(_no_data_df)}개 (클릭해서 보기)"):
            st.dataframe(_no_data_df, use_container_width=True)

    def normalize_url(u):
        if not isinstance(u, str) or not u.strip():
            return None
        u = u.strip()
        if not u.startswith('http://') and not u.startswith('https://'):
            u = 'https://' + u
        return u

    output_df['홈페이지 주소'] = output_df['홈페이지 주소'].apply(normalize_url)

    display_df = output_df.copy()
    for col in ['매출액(억원)', '영업이익(억원)', '당기순이익(억원)']:
        if col in display_df.columns:
            display_df[col] = display_df[col].apply(lambda x: f"{int(x):,}" if pd.notna(x) else "")

    _prev_toggle = st.session_state.get("_select_all_value", False)
    with st.container(key="select_all_container"):
        select_all_toggle = st.checkbox("전체 선택 / 전체 해제", value=_prev_toggle)
    if select_all_toggle != _prev_toggle:
        st.session_state["_select_all_value"] = select_all_toggle
        st.session_state["_editor_key_counter"] = st.session_state.get("_editor_key_counter", 0) + 1
        st.rerun()

    editor_df = display_df.copy()
    editor_df.insert(0, '선택', st.session_state.get("_select_all_value", False))

    edited_df = st.data_editor(
        editor_df,
        use_container_width=True,
        hide_index=True,
        disabled=[c for c in editor_df.columns if c != '선택'],
        column_config={
            "선택": st.column_config.CheckboxColumn("선택", pinned=True),
            "회사명": st.column_config.TextColumn("회사명", pinned=True),
            "대표상품/브랜드": st.column_config.TextColumn("대표상품/브랜드", width="small"),
            "홈페이지 주소": st.column_config.LinkColumn("홈페이지 주소", display_text="바로가기"),
            "관련기사": st.column_config.LinkColumn("관련기사", display_text="기사보기"),
            "기업정보(bizno)": st.column_config.LinkColumn("기업정보(bizno)", display_text="상세보기"),
            "증권": st.column_config.LinkColumn("증권", display_text="정보 보기"),
            "통신판매업조회": st.column_config.LinkColumn("통신판매업조회", display_text="확인하기"),
        },
        key=f"results_editor_{st.session_state.get('_editor_key_counter', 0)}",
    )

    checked = edited_df[edited_df['선택']]
    st.markdown("**🔎 체크한 회사(들)의 업종으로 재검색하기**")
    if not checked.empty:
        picked_industries = sorted(output_df.loc[checked.index, '업종'].unique().tolist())
        st.caption(f"체크한 {len(checked)}개사의 업종: " + ", ".join(picked_industries))
        if st.button(f"이 업종({len(picked_industries)}개)으로 재검색"):
            st.session_state["pending_industry_search"] = picked_industries
            del st.session_state["last_output_df"]
            st.rerun()
    else:
        st.caption("표 왼쪽 '선택' 칸에 체크하면, 그 회사(들)의 업종으로 재검색할 수 있습니다.")

    buf = io.BytesIO()
    export_df = output_df.loc[checked.index] if not checked.empty else output_df
    export_df.to_excel(buf, index=False)
    buf.seek(0)

    # 서식 개선: 링크 컬럼은 보기 좋은 문구+하이퍼링크로, 헤더는 연두색 배경, 값 있는 셀은 전체 테두리
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Border, Side

        wb = load_workbook(buf)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]

        # 1) 링크 컬럼: 원래 주소는 하이퍼링크로 걸고, 셀에는 화면과 같은 짧은 문구만 표시
        link_display_text = {
            "관련기사": "기사보기",
            "기업정보(bizno)": "상세보기",
            "증권": "정보 보기",
            "통신판매업조회": "확인하기",
            "홈페이지 주소": "바로가기",
        }
        for col_name, display_text in link_display_text.items():
            if col_name not in headers:
                continue
            col_idx = headers.index(col_name) + 1
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                url = cell.value
                if url and isinstance(url, str) and url.strip():
                    cell.hyperlink = url
                    cell.value = display_text
                    cell.style = "Hyperlink"

        # 2) 맨 윗줄(헤더) 배경을 연한 초록으로
        header_fill = PatternFill(start_color="D9F2D9", end_color="D9F2D9", fill_type="solid")
        for cell in ws[1]:
            cell.fill = header_fill

        # 3) 값이 있는 셀에는 전부 테두리 적용
        thin_side = Side(style="thin", color="B0B0B0")
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=len(headers)):
            for cell in row:
                if cell.value is not None and str(cell.value).strip() != '':
                    cell.border = thin_border

        buf = io.BytesIO()
        wb.save(buf)
    except Exception:
        buf.seek(0)  # 서식 변환 실패해도 기본 엑셀은 그대로 다운로드되게 둠

    download_label = (
        f"📥 체크한 {len(export_df)}개사만 엑셀로 다운로드"
        if not checked.empty else "📥 전체 엑셀로 다운로드"
    )
    st.download_button(
        download_label,
        data=buf.getvalue(),
        file_name="screener_result.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
